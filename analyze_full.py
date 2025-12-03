#!/usr/bin/env python3
"""Full analysis of the Excel file"""

from openpyxl import load_workbook
import json

wb = load_workbook('zwift.xlsx')
ws = wb.active

# Get header row (row 2)
headers = [cell.value for cell in ws[2]]

# Extract data
routes = []
users = ['John', 'Alex', 'Karen', 'Carol', 'Ali']
date_columns = []

# Find date columns (columns S onwards, index 18+)
for col_idx, header in enumerate(headers):
    if isinstance(header, type(ws[2][col_idx].value)) and hasattr(header, 'year'):
        date_columns.append((col_idx, header))

# Process route rows (starting from row 3)
for row_idx in range(3, ws.max_row + 1):
    row = ws[row_idx]
    map_name = row[0].value
    route_name = row[1].value
    
    if not map_name or not route_name:
        continue
    
    route_data = {
        'map': map_name,
        'route': route_name,
        'length_km': row[2].value,
        'length_miles': row[3].value,
        'elevation': row[4].value,
        'lead_in': row[5].value,
        'badge_xp': row[6].value,
        'users': {}
    }
    
    # Get user selections (columns I-M, index 8-12)
    for i, user in enumerate(users):
        col_idx = 8 + i
        if col_idx < len(row):
            cell_value = row[col_idx].value
            route_data['users'][user] = cell_value == 'x' or cell_value == 'X'
    
    # Get vote winner (column O, index 14)
    if len(row) > 14:
        route_data['vote_winner'] = row[14].value == 'x' or row[14].value == 'X'
    
    routes.append(route_data)

print(f"Total routes: {len(routes)}")
print(f"Users: {users}")
print(f"Date columns: {len(date_columns)}")
print(f"\nFirst 5 routes:")
for route in routes[:5]:
    print(f"  {route['map']} - {route['route']}: {route['length_miles']}")

# Save as JSON for easier processing
with open('routes_data.json', 'w') as f:
    json.dump({
        'goal_km': 500,
        'goal_miles': 312.5,
        'users': users,
        'routes': routes
    }, f, indent=2, default=str)

print("\nData exported to routes_data.json")

