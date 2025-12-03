#!/usr/bin/env python3
"""Analyze the Excel file structure"""

try:
    from openpyxl import load_workbook
    wb = load_workbook('zwift.xlsx')
    print('Sheet names:', wb.sheetnames)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\n=== Sheet: {sheet_name} ===')
        print(f'Dimensions: {ws.dimensions}')
        print(f'Max row: {ws.max_row}, Max column: {ws.max_column}')
        
        print('\nFirst 15 rows:')
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i > 15:
                break
            print(f'Row {i}: {row}')
            
except ImportError:
    print("openpyxl not installed. Trying pandas...")
    try:
        import pandas as pd
        xls = pd.ExcelFile('zwift.xlsx')
        print('Sheet names:', xls.sheet_names)
        
        for sheet_name in xls.sheet_names:
            df = pd.read_excel('zwift.xlsx', sheet_name=sheet_name)
            print(f'\n=== Sheet: {sheet_name} ===')
            print(f'Shape: {df.shape}')
            print(f'Columns: {list(df.columns)}')
            print('\nFirst 10 rows:')
            print(df.head(10).to_string())
    except ImportError:
        print("Neither openpyxl nor pandas installed.")
        print("Please install: pip install openpyxl pandas")

